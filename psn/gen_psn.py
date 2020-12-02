import openpyxl
import os
import sys

# data_only=True, for read value, otherwise read the math equation in cell
fgo_xls  = openpyxl.load_workbook('FGO.xlsx', data_only=True)
fgo_psn  = open('PSN.py', 'w+')

worksheets = fgo_xls.sheetnames
print(worksheets)

# sheet = fgo_xls.get_sheet_by_name('Pexel LUT')
sheet = fgo_xls['Pexel LUT']
# print(sheet)
# print(sheet.title)

# print(fgo_xls.active)
# print(sheet['B1'], sheet['B1'].value)
# print(sheet.cell(row=1, column=2).value)
# print(sheet.max_row, sheet.max_column)

basic_func_name = []  # column=1
basic_func_calx = []  # column=4
basic_func_caly = []  # column=5
basic_func_slpt = []  # column=6
basic_func_skil = []  # column=8

# from row 3 to end
row_start = 3
row_end   = sheet.max_row + 1

# print(row_start, row_end)
# exit()

print(row_start, row_end)

for i in range(row_start, row_end):
    basic_func_name.append(sheet.cell(row=i, column=1).value)

for i in range(row_start, row_end):
    basic_func_calx.append(sheet.cell(row=i, column=4).value)

for i in range(row_start, row_end):
    basic_func_caly.append(sheet.cell(row=i, column=5).value)

for i in range(row_start, row_end):
    basic_func_slpt.append(sheet.cell(row=i, column=6).value)

for i in range(row_start, row_end):
    basic_func_skil.append(sheet.cell(row=i, column=8).value)

print('\n\n')
print('+-----+------------+-----------+------------+---+')
print('| No  | KEY        | POSITION  | DURATION   | F |')
print('+-----+------------+-----------+------------+---+')
for i in range(len(basic_func_name)):
    print('+-----+------------+-----------+------------+---+')
    print('| %3d | %-10s | %4d,%4d | %7d ms | %1d |' % (i, basic_func_name[i], basic_func_calx[i], basic_func_caly[i], basic_func_slpt[i], basic_func_skil[i]))

    if i == len(basic_func_name) - 1:
        print('+-----+------------+-----------+------------+---+')

# generate fgo_psn.js file
header = '''
__metaclass__ = type

import time
from util.ats import tap
from util.global0 import speed


class PSN:
    def __init__(self):
        pass
'''

fgo_psn.write(header)
fgo_psn.write('\n')

for i in range(len(basic_func_name)):
    fgo_psn.write('    def %s(self, duration=None):\n' % basic_func_name[i])
    fgo_psn.write('        if duration is None:\n')
    fgo_psn.write('            duration = %s\n' % str(int(basic_func_slpt[i])/1000))
    fgo_psn.write('        duration = duration * speed()\n')
    fgo_psn.write('        tap(%s, %s)\n' % (str(basic_func_calx[i]), str(basic_func_caly[i])))
    fgo_psn.write('        time.sleep(duration)\n\n')

fgo_psn.close()
























